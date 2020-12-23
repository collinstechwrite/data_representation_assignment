import webbrowser
import os
import json
import requests
from newsapi import NewsApiClient
from openpyxl import Workbook, load_workbook
import collections
import matplotlib.pyplot as plt


from flask import Flask, request, render_template
from flask_mysqldb import MySQL
import io
import base64

from matplotlib.backends.backend_agg import FigureCanvasAgg

import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt


import seaborn as sns


app = Flask(__name__)
app.config['MYSQL_USER'] = 'sql2383132'
app.config['MYSQL_PASSWORD'] = 'qL1!bV7*'
app.config['MYSQL_HOST'] = 'sql2.freemysqlhosting.net'
app.config['MYSQL_DB'] = 'sql2383132'
app.config['MYSQL_CURSORCLASS'] = 'DictCursor'
mysql = MySQL(app)


jdict = {}


@app.route('/visualize')
def visualize():
    import matplotlib
    import matplotlib.pyplot as plt
    import numpy as np



    global jdict
    # this string will be used for doing word count analysis
    my_string_for_word_count = ""
    for row in jdict:

        headline_title = row['title']
        # remove/replace inverted commas to avoid SQL errors when passing data to database
        headline_title = headline_title.replace("'", "")

        url = row['url']

        # adding one healine at a time to the string for word count analysis
        my_string_for_word_count = my_string_for_word_count + headline_title


    #this code below does a word count analysis
    #https://www.w3resource.com/python-exercises/string/python-data-type-string-exercise-12.php
    def word_count(str):
        counts = {}
        words = str.split()

        for word in words:
            if word in counts:
                counts[word] += 1
            else:
                counts[word] = 1

        return counts

    counted_words =  word_count(my_string_for_word_count)


    sorted_dict = collections.OrderedDict(counted_words)
    x = sorted_dict

    #https://stackoverflow.com/questions/613183/how-do-i-sort-a-dictionary-by-value
    #sorted_x = sorted(x.items(), key=lambda kv: kv[1])

    #https://stackoverflow.com/questions/646644/how-to-get-last-items-of-a-list-in-python
    #my_graph_data = sorted_x[-10:]
    #my_graph_data_dict = dict(my_graph_data)
     
    #https://thispointer.com/different-ways-to-remove-a-key-from-dictionary-in-python/
    common_words = ['the','of','and','a','to','in','is','you','that','it','he','was',
    'for','on',
    'are',
    'as',
    'with',
    'his',
    'they',
    'I',
    'at',
    'be',
    'this',
    'have',
    'from',
    'or',
    'one',
    'had',
    'by',
    'word',
    'but',
    'not',
    'what',
    'all',
    'were',
    'we',
    'when',
    'your',
    'can',
    'said',
    'there',
    'use',
    'an',
    'each',
    'which',
    'she',
    'do',
    'how',
    'their',
    'if',
    'will',
    'up',
    'other',
    'about',
    'out',
    'many',
    'then',
    'them',
    'these',
    'so',
    'some',
    'her',
    'would',
    'make',
    'like',
    'him',
    'into',
    'time',
    'has',
    'look',
    'two',
    'more',
    'write',
    'go',
    'see',
    'number',
    'no',
    'way',
    'could',
    'people',
    'my',
    'than',
    'first',
    'been',
    'call',
    'who',
    'its',
    'now',
    'find',
    'long',
    'down',
    'day',
    'did',
    'get',
    'come',
    'made',
    'may',
    'part',
    ]

    for word in common_words:
        try:
            #my_graph_data_dict.pop(word)
            x.pop(word)
        except:
            continue

    #https://www.kite.com/python/answers/how-to-plot-a-bar-chart-using-a-dictionary-in-matplotlib-in-python
    #a_dictionary = x
    #keys = my_graph_data_dict.keys()
    #values = my_graph_data_dict.values()

    keys = x.keys()
    values = x.values()

 
    #https://www.kite.com/python/answers/how-to-rotate-axis-labels-in-matplotlib-in-python
    plt.xticks(rotation=45)
    plt.yticks(rotation=90)
    #https://showmecode.info/matplotlib/bar/change-bar-color/
    plt.bar(keys, values,  color=['red', 'blue', 'purple', 'green', 'lavender'])
    plt.ylabel('Occurences of word')
    plt.title('Most Frequent Words In Headlines')
    #https://www.kite.com/python/answers/how-save-a-matplotlib-plot-as-a-pdf-file-in-python


    #here is the trick save your figure into a bytes object and you can afterwards expose it via flas
    bytes_image = io.BytesIO()
    plt.savefig('img')
    
    return send_file(img,mimetype='img')


    


@app.route('/StyleSheet.css')
def my_style_sheet():
    return """
input[type=text], select {
  width: 100%;
  padding: 12px 20px;
  margin: 8px 0;
  display: inline-block;
  border: 1px solid #ccc;
  border-radius: 4px;
  box-sizing: border-box;
}
input[type=submit] {
  width: 100%;
  background-color: #4CAF50;
  color: white;
  padding: 14px 20px;
  margin: 8px 0;
  border: none;
  border-radius: 4px;
  cursor: pointer;
}
input[type=submit]:hover {
  background-color: #45a049;
}
div {
  border-radius: 5px;
  background-color: #f2f2f2;
  padding: 20px;
}
img {
   padding:1px;
   border:1px solid #021a40;
   background-color:#ff0;
}
body {
  font-family: Helvetica, sans-serif;
}
body {
  --stripe: #cfd8dc;
  --bg: #e1e1e1;
  background: linear-gradient(135deg, var(--bg) 25%, transparent 25%) -50px 0,
    linear-gradient(225deg, var(--bg) 25%, transparent 25%) -50px 0,
    linear-gradient(315deg, var(--bg) 25%, transparent 25%),
    linear-gradient(45deg, var(--bg) 25%, transparent 25%);
  background-size: 100px 100px;
  background-color: var(--stripe);
}
.styled-table {
    border-collapse: collapse;
    margin: 25px 0;
    font-size: 0.9em;
    font-family: sans-serif;
    min-width: 400px;
    box-shadow: 0 0 20px rgba(0, 0, 0, 0.15);
}
.styled-table thead tr {
    background-color: #009879;
    color: #ffffff;
    text-align: left;
}
.styled-table th,
.styled-table td {
    padding: 12px 15px;
}
.styled-table tbody tr {
    border-bottom: 1px solid #dddddd;
}
.styled-table tbody tr:nth-of-type(even) {
    background-color: #f3f3f3;
}
.styled-table tbody tr:last-of-type {
    border-bottom: 2px solid #009879;
}
.styled-table tbody tr.active-row {
    font-weight: bold;
    color: #009879;
}
/* Style the header with a grey background and some padding */
.header {
  overflow: hidden;
  background-color: #f1f1f1;
  padding: 20px 10px;
}
/* Style the header links */
.header a {
  float: left;
  color: black;
  text-align: center;
  padding: 12px;
  text-decoration: none;
  font-size: 18px;
  line-height: 25px;
  border-radius: 4px;
}
/* Style the logo link (notice that we set the same value of line-height and font-size to prevent the header to increase when the font gets bigger */
.header a.logo {
  font-size: 25px;
  font-weight: bold;
}
/* Change the background color on mouse-over */
.header a:hover {
  background-color: #ddd;
  color: black;
}
/* Style the active/current link*/
.header a.active {
  background-color: dodgerblue;
  color: white;
}
/* Float the link section to the right */
.header-right {
  float: right;
}
/* Add media queries for responsiveness - when the screen is 500px wide or less, stack the links on top of each other */
@media screen and (max-width: 500px) {
  .header a {
    float: none;
    display: block;
    text-align: left;
  }
  .header-right {
    float: none;
  }
}
"""











@app.route('/')
def my_form():
    return render_template('my-form.html')

@app.route('/', methods=['POST'])
def my_form_post():

   
    def generate_graph():
        global jdict
        # this string will be used for doing word count analysis
        my_string_for_word_count = ""
        for row in jdict:

            headline_title = row['title']
            # remove/replace inverted commas to avoid SQL errors when passing data to database
            headline_title = headline_title.replace("'", "")

            url = row['url']

            # adding one healine at a time to the string for word count analysis
            my_string_for_word_count = my_string_for_word_count + headline_title


        #this code below does a word count analysis
        #https://www.w3resource.com/python-exercises/string/python-data-type-string-exercise-12.php
        def word_count(str):
            counts = {}
            words = str.split()

            for word in words:
                if word in counts:
                    counts[word] += 1
                else:
                    counts[word] = 1

            return counts

        counted_words =  word_count(my_string_for_word_count)


        sorted_dict = collections.OrderedDict(counted_words)
        x = sorted_dict

        #https://stackoverflow.com/questions/613183/how-do-i-sort-a-dictionary-by-value
        #sorted_x = sorted(x.items(), key=lambda kv: kv[1])

        #https://stackoverflow.com/questions/646644/how-to-get-last-items-of-a-list-in-python
        #my_graph_data = sorted_x[-10:]
        #my_graph_data_dict = dict(my_graph_data)
         
        #https://thispointer.com/different-ways-to-remove-a-key-from-dictionary-in-python/
        common_words = ['the','of','and','a','to','in','is','you','that','it','he','was',
        'for','on',
        'are',
        'as',
        'with',
        'his',
        'they',
        'I',
        'at',
        'be',
        'this',
        'have',
        'from',
        'or',
        'one',
        'had',
        'by',
        'word',
        'but',
        'not',
        'what',
        'all',
        'were',
        'we',
        'when',
        'your',
        'can',
        'said',
        'there',
        'use',
        'an',
        'each',
        'which',
        'she',
        'do',
        'how',
        'their',
        'if',
        'will',
        'up',
        'other',
        'about',
        'out',
        'many',
        'then',
        'them',
        'these',
        'so',
        'some',
        'her',
        'would',
        'make',
        'like',
        'him',
        'into',
        'time',
        'has',
        'look',
        'two',
        'more',
        'write',
        'go',
        'see',
        'number',
        'no',
        'way',
        'could',
        'people',
        'my',
        'than',
        'first',
        'been',
        'call',
        'who',
        'its',
        'now',
        'find',
        'long',
        'down',
        'day',
        'did',
        'get',
        'come',
        'made',
        'may',
        'part',
        ]

        for word in common_words:
            try:
                #my_graph_data_dict.pop(word)
                x.pop(word)
            except:
                continue

        #https://www.kite.com/python/answers/how-to-plot-a-bar-chart-using-a-dictionary-in-matplotlib-in-python
        #a_dictionary = x
        #keys = my_graph_data_dict.keys()
        #values = my_graph_data_dict.values()

        keys = x.keys()
        values = x.values()


        #https://www.kite.com/python/answers/how-to-rotate-axis-labels-in-matplotlib-in-python
        plt.xticks(rotation=45)
        plt.yticks(rotation=90)
        #https://showmecode.info/matplotlib/bar/change-bar-color/
        plt.bar(keys, values,  color=['red', 'blue', 'purple', 'green', 'lavender'])
        plt.ylabel('Occurences of word')
        plt.title('Most Frequent Words In Headlines')
        #https://www.kite.com/python/answers/how-save-a-matplotlib-plot-as-a-pdf-file-in-python
        plt.savefig("static/plot.png")
    def receive_text_from_form(text):
        global jdict
        
        # This is a very busy function, which handles extracting news from API, creating and saving to spreadsheet, saving records to MYSQL Server, doing word count analysis
        # creating and saving a matplotlib graph

        
        #Setting up extractions from News API

        news_keyword = text

        # https://newsapi.org/docs/client-libraries/python
        newsapi = NewsApiClient(api_key='0fb13acc3bc8480eafedb87afa941f7e')


        # /v2/everything
        data = newsapi.get_everything(q=news_keyword)

        jdict = data.get('articles')
        #generate_graph(jdict)



        #Creating spreadsheet

        Row_Incrementer = 1

        workbook = Workbook()
        sheet = workbook.active
        sheet["A" + str(Row_Incrementer)] = "Headline Title"
        sheet["B" + str(Row_Incrementer)] = "URL"
        workbook.save(filename="analysis.xlsx")
        workbook.close()

        
        #Creating list to return to web page
        
        my_list = []
        my_string_for_word_count = "" #this string will be used for doing word count analysis

        
        for row in jdict:
            #Setting up spreadsheet connection
            Row_Incrementer += 1 #incrementing the row number to add next record below
            workbook = load_workbook(filename="analysis.xlsx")
            sheet = workbook.active
            #Setting up database connection
            cur = mysql.connection.cursor()
            #Setting up extraction of data to list which will be published to web page
            transfer = row['title'],row['url']
            headline = row['title']
            headline = headline.replace("'", "") #remove/replace inverted commas to avoid SQL errors when passing data to database
            url = row['url']
            my_list.append(transfer)
            #Setting up extraction of data to online mysql database
            SQL = ("INSERT INTO headlinetitles (Title, Url) values ('"+ headline +"','"+ url +"');")
            cur.execute(SQL)
            mysql.connection.commit()
            #Setting up extraction of data to spreadsheet
            sheet["A" + str(Row_Incrementer)] = headline
            sheet["B" + str(Row_Incrementer)] = url
            workbook.save(filename="analysis.xlsx")
            workbook.close()
        return my_list


    
    text = request.form['text']


    #get current working directory
    path = os.getcwd()

    
    list = receive_text_from_form(text) #this updates the global variable jdict with extracted newspaper headlines

    generate_graph() #accesses the global variable jdict to produce a graph
    filelocation = "Find yous files here", path, "/analysis.xlsx"
    
    return render_template("my-form.html", list=list, filelocation=filelocation)



@app.route('/delete_database')
def delete_database():


   #Port number: 3306

    cur = mysql.connection.cursor()
    cur.execute('''DELETE FROM headlinetitles;''')
    mysql.connection.commit()
    results = cur.fetchall()
    results = str(results)
    return """
    <link href="StyleSheet.css" rel="stylesheet">
    <table class="styled-table"><thead><tr><th><centre>Database Contents On MySQL Server</centre></th></tr></thead><tbody><tr class="active-row"><td>""" + results + """ </td></tr></tbody></tr></table>
    <img src="static/images/plot.png" alt="A Matplotlibplot">
    <button onclick="goBack()">Go Back</button>
<script>
function goBack() {
  window.history.back();
}
</script>
    """ 

@app.route('/database')




def database():

   #Port number: 3306

    cur = mysql.connection.cursor()
    cur.execute('''SELECT * FROM headlinetitles''')
    results = cur.fetchall()
    results = str(results)
    return """ 
    <link href="StyleSheet.css" rel="stylesheet">
  
    <center>
    <table class="styled-table"><thead><tr><th><centre>Database Contents On MySQL Server</centre></th></tr></thead><tbody><tr class="active-row"><td>""" + results + """ </td></tr></tbody></tr></table>
    <img src="static/images/plot.png" alt="A Matplotlibplot">
    <a href="/delete_database">Click here to delete database content</a>
    <button onclick="goBack()">Go Back</button>
<script>
function goBack() {
  window.history.back();
}
</script>
</center>
    """ 



if __name__ == "__main__":
    app.run()
